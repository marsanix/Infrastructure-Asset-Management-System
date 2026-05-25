"""
Report Management endpoint.

Security:
- RBAC: hanya user dengan permission report:read
- CSV export: sanitasi formula injection (defusedcsv / manual prefix stripping)
- Excel export: openpyxl, cell value di-cast ke string/number — tidak ada formula
- Response header Content-Disposition dengan filename yang di-sanitize
- Tidak ada raw SQL — semua via SQLAlchemy ORM + view
- Pagination untuk JSON response (max 1000 rows per request)
"""
import io
import re
from datetime import date
from flask import request, jsonify, send_file, current_app
from flask_jwt_extended import get_jwt_identity

from . import api_v1
from ...extensions import db
from ...models.asset import Asset, AssetStatus
from ...models.model_device import DeviceModel
from ...models.brand import Brand
from ...models.category import Category
from ...models.location import Location
from ...models.employee import Employee
from ...models.department import Department
from ...models.network_detail import NetworkDetail
from ...utils.rbac import require_permission

# Karakter yang memicu formula injection di spreadsheet
_FORMULA_PREFIXES = ('=', '+', '-', '@', '\t', '\r')


def _sanitize_cell(value) -> str:
    """
    Sanitasi nilai cell untuk CSV/Excel export.
    Prefix formula injection characters dengan single quote.
    Referensi: OWASP CSV Injection prevention.
    """
    if value is None:
        return ''
    s = str(value)
    if s.startswith(_FORMULA_PREFIXES):
        return "'" + s   # prefix dengan ' mencegah Excel menginterpretasi sebagai formula
    return s


def _build_asset_query():
    """Query lengkap aset dengan semua relasi."""
    return (
        db.session.query(
            Asset, DeviceModel, Brand, Category, Location,
            Employee, Department, NetworkDetail
        )
        .join(DeviceModel,   Asset.model_id    == DeviceModel.id)
        .join(Brand,         DeviceModel.brand_id    == Brand.id,    isouter=True)
        .join(Category,      DeviceModel.category_id == Category.id, isouter=True)
        .join(Location,      Asset.location_id  == Location.id)
        .outerjoin(Employee,   Asset.employee_id  == Employee.id)
        .outerjoin(Department, Employee.department_id == Department.id)
        .outerjoin(NetworkDetail, Asset.id == NetworkDetail.asset_id)
        .filter(Asset.deleted_at.is_(None))
    )


def _apply_filters(query, args):
    """Terapkan filter dari query params — semua via ORM."""
    status      = args.get('status', '').strip()
    category_id = args.get('category_id', type=int)
    location_id = args.get('location_id', type=int)
    search      = args.get('search', '').strip()[:100]

    if status:
        valid = [s.value for s in AssetStatus]
        if status in valid:
            query = query.filter(Asset.status == status)

    if category_id:
        query = query.filter(DeviceModel.category_id == category_id)

    if location_id:
        query = query.filter(Asset.location_id == location_id)

    if search:
        like = f'%{search}%'
        query = query.filter(
            db.or_(Asset.asset_tag.ilike(like), Asset.serial_number.ilike(like))
        )

    return query.order_by(Category.name, Asset.asset_tag)


def _row_to_dict(row) -> dict:
    """Konversi row query ke dict — explicit field list."""
    asset, model, brand, category, location, employee, department, network = row

    # Hitung sisa garansi
    warranty_remaining = None
    if asset.purchase_date and asset.warranty_months:
        from dateutil.relativedelta import relativedelta
        expiry = asset.purchase_date + relativedelta(months=asset.warranty_months)
        today  = date.today()
        if expiry > today:
            delta = relativedelta(expiry, today)
            warranty_remaining = delta.months + delta.years * 12
        else:
            warranty_remaining = 0

    return {
        'asset_tag':          asset.asset_tag,
        'serial_number':      asset.serial_number,
        'po_number':          asset.po_number or '',
        'category':           category.name if category else '',
        'brand':              brand.name    if brand    else '',
        'model':              model.name,
        'status':             asset.status.value,
        'location':           location.name,
        'employee':           employee.name   if employee   else 'IT Inventory / Server',
        'department':         department.name if department else 'Infrastructure',
        'ip_address':         str(network.ip_address)  if network and network.ip_address  else '',
        'mac_address':        str(network.mac_address) if network and network.mac_address else '',
        'purchase_date':      asset.purchase_date.isoformat() if asset.purchase_date else '',
        'warranty_months':    asset.warranty_months or '',
        'warranty_remaining': warranty_remaining if warranty_remaining is not None else '',
        'os_license':         asset.os_license or '',
    }


REPORT_HEADERS = [
    'Asset Tag', 'Serial Number', 'PO Number', 'Category', 'Brand', 'Model',
    'Status', 'Location', 'Employee', 'Department', 'IP Address', 'MAC Address',
    'Purchase Date', 'Warranty (months)', 'Warranty Remaining (months)', 'OS License',
]

REPORT_KEYS = [
    'asset_tag', 'serial_number', 'po_number', 'category', 'brand', 'model',
    'status', 'location', 'employee', 'department', 'ip_address', 'mac_address',
    'purchase_date', 'warranty_months', 'warranty_remaining', 'os_license',
]


# ── JSON Report ───────────────────────────────────────────────

@api_v1.get('/reports/assets')
@require_permission('report:read')
def asset_report_json():
    try:
        page     = max(1, int(request.args.get('page', 1)))
        per_page = min(max(1, int(request.args.get('per_page', 50))), 1000)
    except (ValueError, TypeError):
        return jsonify({'error': 'Invalid pagination parameters'}), 400

    query      = _build_asset_query()
    query      = _apply_filters(query, request.args)
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)

    return jsonify({
        'data':     [_row_to_dict(row) for row in pagination.items],
        'total':    pagination.total,
        'page':     pagination.page,
        'per_page': pagination.per_page,
        'pages':    pagination.pages,
    }), 200


# ── Status Summary ────────────────────────────────────────────

@api_v1.get('/reports/assets/summary')
@require_permission('report:read')
def asset_summary():
    rows = (
        db.session.query(Asset.status, db.func.count(Asset.id))
        .filter(Asset.deleted_at.is_(None))
        .group_by(Asset.status)
        .all()
    )
    return jsonify({
        'data': [{'status': r[0].value, 'total': r[1]} for r in rows]
    }), 200


# ── CSV Export ────────────────────────────────────────────────

@api_v1.get('/reports/assets/export/csv')
@require_permission('report:read')
def export_csv():
    import csv

    query = _build_asset_query()
    query = _apply_filters(query, request.args)
    rows  = query.limit(10000).all()   # hard limit 10k rows

    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_ALL)

    # Header row
    writer.writerow(REPORT_HEADERS)

    # Data rows — sanitasi setiap cell (formula injection prevention)
    for row in rows:
        d = _row_to_dict(row)
        writer.writerow([_sanitize_cell(d[k]) for k in REPORT_KEYS])

    output.seek(0)
    filename = f'asset_report_{date.today().isoformat()}.csv'

    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),  # BOM untuk Excel compatibility
        mimetype='text/csv',
        as_attachment=True,
        download_name=_sanitize_filename(filename),
    )


# ── Excel Export ──────────────────────────────────────────────

@api_v1.get('/reports/assets/export/excel')
@require_permission('report:read')
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    query = _build_asset_query()
    query = _apply_filters(query, request.args)
    rows  = query.limit(10000).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Asset Report'

    # Header styling
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='0F62FE')  # IBM Blue

    for col_idx, header in enumerate(REPORT_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Data rows — nilai di-cast ke tipe yang tepat, tidak ada formula
    for row_idx, row in enumerate(rows, 2):
        d = _row_to_dict(row)
        for col_idx, key in enumerate(REPORT_KEYS, 1):
            raw = d[key]
            # Cast ke int/float jika numerik, string jika tidak
            if isinstance(raw, int):
                value = raw
            elif isinstance(raw, str) and raw.isdigit():
                value = int(raw)
            else:
                # Sanitasi formula injection — openpyxl tidak execute formula
                # tapi kita tetap strip prefix untuk keamanan
                value = _sanitize_cell(raw)
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-width kolom
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'asset_report_{date.today().isoformat()}.xlsx'

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=_sanitize_filename(filename),
    )


def _sanitize_filename(filename: str) -> str:
    """Sanitasi nama file download — cegah path traversal."""
    # Hanya izinkan alphanumeric, dash, underscore, dot
    safe = re.sub(r'[^A-Za-z0-9_\-\.]', '_', filename)
    return safe[:100]
